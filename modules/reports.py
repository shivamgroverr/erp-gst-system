# modules/reports.py
from flask import Blueprint, render_template, request, session,                   jsonify, Response, flash, redirect, url_for
from flask_login import login_required, current_user
from extensions import db
from models import (Account, JournalLine, JournalHeader, Bill,
                    BillItem, Party, Item, StockLedger, FixedAsset,
                    TdsEntry, ComplianceAlert, GstReturn)
from sqlalchemy import func, extract, or_, and_, text
from datetime import date, datetime
from io import BytesIO
import json, openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch

reports_bp = Blueprint("reports", __name__)

# ═══════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════
def fy_dates(fy):
    """Return (start_date, end_date) for FY like '2025-26'"""
    y1 = int(fy[:4])
    return date(y1, 4, 1), date(y1+1, 3, 31)

def xl_header(ws, headers, row=1, fill="1F4E79", font_color="FFFFFF"):
    from openpyxl.styles import Font, Alignment, PatternFill
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = Font(bold=True, color=font_color)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal="center")

def xl_title(ws, title, cols=10, row=1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=title)
    c.font = Font(bold=True, size=14)
    c.alignment = Alignment(horizontal="center")

def money(v): return round(float(v or 0), 2)

# ═══════════════════════════════════════════
#  1. REPORTS HUB DASHBOARD
# ═══════════════════════════════════════════
@reports_bp.route("/reports")
@login_required
def hub():
    cid    = session.get("company_id")
    fy     = session.get("fin_year")
    period = request.args.get("period", date.today().strftime("%m-%Y"))

    # Quick summary numbers for the cards
    sales_total = db.session.query(func.sum(Bill.total_amount)).filter_by(
        company_id=cid, fin_year=fy, bill_type="Sales", is_cancelled=False
    ).scalar() or 0

    purchase_total = db.session.query(func.sum(Bill.total_amount)).filter_by(
        company_id=cid, fin_year=fy, bill_type="Purchase", is_cancelled=False
    ).scalar() or 0

    pending_alerts = ComplianceAlert.query.filter_by(
        company_id=cid, status="pending"
    ).count()

    pending_gst_returns = GstReturn.query.filter_by(
        company_id=cid, status="pending"
    ).count() if hasattr(GstReturn, "__tablename__") else 0

    return render_template("reports/hub.html",
        period=period, fy=fy,
        sales_total=money(sales_total),
        purchase_total=money(purchase_total),
        pending_alerts=pending_alerts,
        pending_gst_returns=pending_gst_returns)

# ═══════════════════════════════════════════
#  2. LEDGER
# ═══════════════════════════════════════════
@reports_bp.route("/reports/ledger")
@login_required
def ledger():
    cid       = session.get("company_id")
    fy        = session.get("fin_year")
    acc_id    = request.args.get("acc_id", type=int)
    from_date = request.args.get("from_date")
    to_date   = request.args.get("to_date")
    export    = request.args.get("export","")
    q_text    = request.args.get("q","")

    start, end = fy_dates(fy)
    if from_date:
        try: start = datetime.strptime(from_date, "%Y-%m-%d").date()
        except: pass
    if to_date:
        try: end = datetime.strptime(to_date, "%Y-%m-%d").date()
        except: pass

    accounts = Account.query.filter_by(company_id=cid, is_active=True
               ).order_by(Account.name).all()

    lines = []
    account = None
    opening_dr = opening_cr = 0.0

    if acc_id:
        account = Account.query.get(acc_id)
        # Opening balance (before from_date)
        ob = db.session.query(
            func.sum(JournalLine.debit).label("dr"),
            func.sum(JournalLine.credit).label("cr"),
        ).join(JournalHeader).filter(
            JournalLine.account_id == acc_id,
            JournalHeader.voucher_date < start,
            JournalHeader.company_id  == cid,
            JournalHeader.is_cancelled == False,
        ).first()
        opening_dr = money(ob.dr)
        opening_cr = money(ob.cr)

        # Transactions
        q = db.session.query(JournalLine, JournalHeader).join(JournalHeader).filter(
            JournalLine.account_id   == acc_id,
            JournalHeader.company_id == cid,
            JournalHeader.voucher_date >= start,
            JournalHeader.voucher_date <= end,
            JournalHeader.is_cancelled == False,
        ).order_by(JournalHeader.voucher_date, JournalHeader.id)

        if q_text:
            q = q.filter(or_(
                JournalHeader.narration.ilike(f"%{q_text}%"),
                JournalHeader.voucher_no.ilike(f"%{q_text}%"),
            ))

        balance = opening_dr - opening_cr
        for jl, jh in q.all():
            balance += money(jl.debit) - money(jl.credit)
            lines.append({
                "date"       : jh.voucher_date,
                "voucher_no" : jh.voucher_no,
                "type"       : jh.voucher_type,
                "narration"  : jh.narration,
                "dr"         : money(jl.debit),
                "cr"         : money(jl.credit),
                "balance"    : round(balance, 2),
                "jh_id"      : jh.id,
            })

    if export == "excel":
        return export_ledger_excel(account, lines, opening_dr, opening_cr, fy, start, end)
    if export == "pdf":
        return export_ledger_pdf(account, lines, opening_dr, opening_cr, fy, start, end)

    return render_template("reports/ledger.html",
        accounts=accounts, account=account, lines=lines,
        opening_dr=opening_dr, opening_cr=opening_cr,
        from_date=start, to_date=end, q=q_text,
        acc_id=acc_id)

# ═══════════════════════════════════════════
#  3. TRIAL BALANCE
# ═══════════════════════════════════════════
@reports_bp.route("/reports/trial-balance")
@login_required
def trial_balance():
    cid    = session.get("company_id")
    fy     = session.get("fin_year")
    export = request.args.get("export","")
    start, end = fy_dates(fy)

    rows = db.session.query(
        Account.id, Account.name, Account.group_name,
        func.sum(JournalLine.debit).label("dr"),
        func.sum(JournalLine.credit).label("cr"),
    ).join(JournalLine, JournalLine.account_id == Account.id
    ).join(JournalHeader, JournalHeader.id == JournalLine.journal_header_id
    ).filter(
        Account.company_id == cid,
        JournalHeader.company_id == cid,
        JournalHeader.voucher_date.between(start, end),
        JournalHeader.is_cancelled == False,
    ).group_by(Account.id, Account.name, Account.group_name
    ).order_by(Account.group_name, Account.name).all()

    data = []
    total_dr = total_cr = 0
    for r in rows:
        dr = money(r.dr); cr = money(r.cr)
        net = dr - cr
        data.append({
            "id": r.id, "name": r.name, "group": r.group_name,
            "dr": dr, "cr": cr,
            "net_dr": max(net, 0), "net_cr": max(-net, 0),
        })
        total_dr += max(net, 0)
        total_cr += max(-net, 0)

    if export == "excel": return export_tb_excel(data, total_dr, total_cr, fy)
    if export == "pdf":   return export_tb_pdf(data, total_dr, total_cr, fy)

    return render_template("reports/trial_balance.html",
        data=data, total_dr=round(total_dr,2),
        total_cr=round(total_cr,2), fy=fy)

# ═══════════════════════════════════════════
#  4. PROFIT & LOSS
# ═══════════════════════════════════════════
@reports_bp.route("/reports/profit-loss")
@login_required
def profit_loss():
    cid    = session.get("company_id")
    fy     = session.get("fin_year")
    export = request.args.get("export","")
    start, end = fy_dates(fy)

    INCOME_GROUPS  = ["Sales","Direct Income","Indirect Income","Other Income"]
    EXPENSE_GROUPS = ["Purchase","Direct Expense","Indirect Expense",
                      "Depreciation","Cost of Goods Sold"]

    def group_totals(groups):
        rows = db.session.query(
            Account.group_name,
            func.sum(JournalLine.debit).label("dr"),
            func.sum(JournalLine.credit).label("cr"),
        ).join(JournalLine, JournalLine.account_id == Account.id
        ).join(JournalHeader, JournalHeader.id == JournalLine.journal_header_id
        ).filter(
            Account.company_id == cid,
            Account.group_name.in_(groups),
            JournalHeader.company_id == cid,
            JournalHeader.voucher_date.between(start, end),
            JournalHeader.is_cancelled == False,
        ).group_by(Account.group_name).all()
        return {r.group_name: money(r.cr) - money(r.dr) for r in rows}

    income_data  = group_totals(INCOME_GROUPS)
    expense_data = group_totals(EXPENSE_GROUPS)

    total_income  = sum(income_data.values())
    total_expense = sum(expense_data.values())
    net_profit    = round(total_income - total_expense, 2)

    if export == "excel": return export_pl_excel(income_data, expense_data, total_income, total_expense, net_profit, fy)
    if export == "pdf":   return export_pl_pdf(income_data, expense_data, total_income, total_expense, net_profit, fy)

    return render_template("reports/profit_loss.html",
        income_data=income_data, expense_data=expense_data,
        total_income=total_income, total_expense=total_expense,
        net_profit=net_profit, fy=fy)

# ═══════════════════════════════════════════
#  5. BALANCE SHEET
# ═══════════════════════════════════════════
@reports_bp.route("/reports/balance-sheet")
@login_required
def balance_sheet():
    cid    = session.get("company_id")
    fy     = session.get("fin_year")
    export = request.args.get("export","")
    start, end = fy_dates(fy)

    ASSET_GROUPS     = ["Fixed Assets","Current Assets","Bank","Cash",
                        "Loans & Advances","Sundry Debtors","Stock in Hand"]
    LIABILITY_GROUPS = ["Capital Account","Reserves & Surplus",
                        "Secured Loans","Unsecured Loans","Current Liabilities",
                        "Sundry Creditors","Duties & Taxes"]

    def grp(groups):
        rows = db.session.query(
            Account.group_name,
            func.sum(JournalLine.debit).label("dr"),
            func.sum(JournalLine.credit).label("cr"),
        ).join(JournalLine, JournalLine.account_id == Account.id
        ).join(JournalHeader, JournalHeader.id == JournalLine.journal_header_id
        ).filter(
            Account.company_id == cid,
            Account.group_name.in_(groups),
            JournalHeader.company_id == cid,
            JournalHeader.voucher_date <= end,
            JournalHeader.is_cancelled == False,
        ).group_by(Account.group_name).all()
        return {r.group_name: money(r.dr) - money(r.cr) for r in rows}

    assets      = grp(ASSET_GROUPS)
    liabilities = grp(LIABILITY_GROUPS)

    total_assets      = sum(assets.values())
    total_liabilities = sum(liabilities.values())

    if export == "excel": return export_bs_excel(assets, liabilities, total_assets, total_liabilities, fy)
    if export == "pdf":   return export_bs_pdf(assets, liabilities, total_assets, total_liabilities, fy)

    return render_template("reports/balance_sheet.html",
        assets=assets, liabilities=liabilities,
        total_assets=round(total_assets,2),
        total_liabilities=round(total_liabilities,2),
        fy=fy)

# ═══════════════════════════════════════════
#  6. CASH FLOW
# ═══════════════════════════════════════════
@reports_bp.route("/reports/cash-flow")
@login_required
def cash_flow():
    cid = session.get("company_id")
    fy  = session.get("fin_year")
    export = request.args.get("export","")
    start, end = fy_dates(fy)

    def cash_movement(group_name):
        r = db.session.query(
            func.sum(JournalLine.debit).label("dr"),
            func.sum(JournalLine.credit).label("cr"),
        ).join(JournalHeader).join(Account,Account.id==JournalLine.account_id
        ).filter(
            Account.company_id == cid,
            Account.group_name == group_name,
            JournalHeader.voucher_date.between(start, end),
            JournalHeader.is_cancelled == False,
        ).first()
        return money(r.dr) - money(r.cr)

    operating = cash_movement("Sundry Debtors") - cash_movement("Sundry Creditors")
    investing  = -cash_movement("Fixed Assets")
    financing  = cash_movement("Capital Account") + cash_movement("Secured Loans")
    net_cash   = round(operating + investing + financing, 2)

    if export == "excel": return export_cashflow_excel(operating, investing, financing, net_cash, fy)
    if export == "pdf":   return export_cashflow_pdf(operating, investing, financing, net_cash, fy)

    return render_template("reports/cash_flow.html",
        operating=operating, investing=investing,
        financing=financing, net_cash=net_cash, fy=fy)

# ═══════════════════════════════════════════
#  7. STOCK SUMMARY
# ═══════════════════════════════════════════
@reports_bp.route("/reports/stock")
@login_required
def stock_summary():
    cid    = session.get("company_id")
    fy     = session.get("fin_year")
    export = request.args.get("export","")
    q_text = request.args.get("q","")

    rows = db.session.query(
        Item.id, Item.name, Item.unit,
        func.sum(StockLedger.in_qty).label("in_qty"),
        func.sum(StockLedger.out_qty).label("out_qty"),
    ).join(StockLedger, StockLedger.item_id == Item.id
    ).filter(
        Item.company_id == cid,
        StockLedger.fin_year == fy,
    ).group_by(Item.id, Item.name, Item.unit).order_by(Item.name).all()

    data = []
    for r in rows:
        in_qty = money(r.in_qty); out_qty = money(r.out_qty)
        closing = in_qty - out_qty
        data.append({
            "id": r.id, "name": r.name, "unit": r.unit,
            "in_qty": in_qty, "out_qty": out_qty, "closing": closing,
            "status": "Low" if closing < 5 else "OK" if closing < 50 else "High",
        })

    if q_text:
        data = [r for r in data if q_text.lower() in r["name"].lower()]
    if export == "excel": return export_stock_excel(data, fy)
    if export == "pdf":   return export_stock_pdf(data, fy)

    return render_template("reports/stock_summary.html", data=data, fy=fy, q=q_text)

# ═══════════════════════════════════════════
#  8. PARTY OUTSTANDING (Debtors / Creditors)
# ═══════════════════════════════════════════
@reports_bp.route("/reports/outstanding")
@login_required
def outstanding():
    cid         = session.get("company_id")
    fy          = session.get("fin_year")
    party_type  = request.args.get("type","Debtor")
    export      = request.args.get("export","")
    start, end  = fy_dates(fy)

    rows = db.session.query(
        Party.id, Party.name, Party.gstin, Party.phone,
        func.sum(Bill.total_amount).label("billed"),
        func.sum(Bill.paid_amount).label("paid"),
    ).join(Bill, Bill.party_id == Party.id
    ).filter(
        Party.company_id == cid,
        Party.party_type == party_type,
        Bill.fin_year    == fy,
        Bill.is_cancelled == False,
    ).group_by(Party.id, Party.name, Party.gstin, Party.phone
    ).order_by(func.sum(Bill.total_amount - Bill.paid_amount).desc()).all()

    data = []
    for r in rows:
        billed = money(r.billed); paid = money(r.paid)
        pending = round(billed - paid, 2)
        if pending > 0.01:
            data.append({
                "id": r.id, "name": r.name, "gstin": r.gstin,
                "phone": r.phone, "billed": billed, "paid": paid,
                "pending": pending,
            })

    if export == "excel": return export_outstanding_excel(data, party_type, fy)
    if export == "pdf":   return export_outstanding_pdf(data, party_type, fy)

    return render_template("reports/outstanding.html",
        data=data, party_type=party_type, fy=fy)

# ═══════════════════════════════════════════
#  9. AUDIT TRAIL
# ═══════════════════════════════════════════
@reports_bp.route("/reports/audit")
@login_required
def audit_trail():
    from models import AuditTrail
    cid    = session.get("company_id")
    q_text = request.args.get("q","")
    export = request.args.get("export","")
    page   = request.args.get("page", 1, type=int)

    query = AuditTrail.query.filter_by(company_id=cid)
    if q_text:
        query = query.filter(or_(
            AuditTrail.action.ilike(f"%{q_text}%"),
            AuditTrail.model_name.ilike(f"%{q_text}%"),
        ))
    query = query.order_by(AuditTrail.created_at.desc())
    pagination = query.paginate(page=page, per_page=50)

    if export == "excel": return export_audit_excel(query.limit(5000).all())
    return render_template("reports/audit.html",
        pagination=pagination, q=q_text)

# ═══════════════════════════════════════════
#  10. COMPLIANCE DASHBOARD
# ═══════════════════════════════════════════
@reports_bp.route("/reports/compliance")
@login_required
def compliance_dashboard():
    cid = session.get("company_id")
    fy  = session.get("fin_year")

    alerts = ComplianceAlert.query.filter_by(
        company_id=cid, status="pending"
    ).order_by(ComplianceAlert.due_date).limit(20).all()

    overdue = [a for a in alerts if a.due_date < date.today()]
    upcoming = [a for a in alerts if a.due_date >= date.today()]

    return render_template("reports/compliance.html",
        overdue=overdue, upcoming=upcoming,
        overdue_count=len(overdue), upcoming_count=len(upcoming))

# ═══════════════════════════════════════════
#  EXCEL EXPORT FUNCTIONS
# ═══════════════════════════════════════════
def _make_response_xlsx(wb, filename):
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment;filename={filename}"})

def _make_response_pdf(buffer, filename):
    buffer.seek(0)
    return Response(buffer,
        mimetype="application/pdf",
        headers={"Content-Disposition": f"attachment;filename={filename}"})

def export_ledger_excel(account, lines, opening_dr, opening_cr, fy, start, end):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ledger"
    xl_title(ws, f"LEDGER — {account.name if account else ''} ({fy})", 7)
    xl_header(ws, ["Date","Voucher No","Type","Narration","Dr","Cr","Balance"], 2)
    r = 3
    ws.append(["Opening Balance","","","",opening_dr, opening_cr,
                round(opening_dr-opening_cr,2)])
    r+=1
    for ln in lines:
        ws.append([ln["date"], ln["voucher_no"], ln["type"],
                   ln["narration"], ln["dr"], ln["cr"], ln["balance"]])
        r+=1
    return _make_response_xlsx(wb, f"Ledger_{account.name if account else 'All'}.xlsx")

def export_tb_excel(data, total_dr, total_cr, fy):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Trial Balance"
    xl_title(ws, f"TRIAL BALANCE — {fy}", 6)
    xl_header(ws, ["Account Name","Group","Dr (₹)","Cr (₹)","Net Dr","Net Cr"], 2)
    r=3
    for d in data:
        ws.append([d["name"],d["group"],d["dr"],d["cr"],d["net_dr"],d["net_cr"]]); r+=1
    ws.append(["TOTAL","",total_dr,total_cr,total_dr,total_cr])
    ws.cell(row=r, column=1).font = Font(bold=True)
    return _make_response_xlsx(wb, f"TrialBalance_{fy}.xlsx")

def export_pl_excel(income_data, expense_data, total_income, total_expense, net_profit, fy):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Profit & Loss"
    xl_title(ws, f"PROFIT & LOSS — {fy}", 4)
    xl_header(ws, ["Head","Income (₹)","Expense (₹)","Net (₹)"], 2)
    r=3
    all_groups = list(set(list(income_data.keys())+list(expense_data.keys())))
    for g in all_groups:
        inc = income_data.get(g,0); exp = expense_data.get(g,0)
        ws.append([g, inc if inc else "", exp if exp else "", inc-exp]); r+=1
    ws.append(["TOTAL", total_income, total_expense, net_profit])
    ws.cell(row=r, column=1).font = Font(bold=True)
    return _make_response_xlsx(wb, f"ProfitLoss_{fy}.xlsx")

def export_bs_excel(assets, liabilities, total_assets, total_liabilities, fy):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Balance Sheet"
    xl_title(ws, f"BALANCE SHEET — {fy}", 4)
    xl_header(ws, ["Liabilities","Amount (₹)","Assets","Amount (₹)"], 2)
    r=3
    lib_rows = list(liabilities.items()); ast_rows = list(assets.items())
    max_rows = max(len(lib_rows), len(ast_rows))
    for i in range(max_rows):
        lib = lib_rows[i] if i < len(lib_rows) else ("","")
        ast = ast_rows[i] if i < len(ast_rows) else ("","")
        ws.append([lib[0], lib[1], ast[0], ast[1]]); r+=1
    ws.append(["TOTAL", total_liabilities, "TOTAL", total_assets])
    ws.cell(row=r, column=1).font = Font(bold=True)
    ws.cell(row=r, column=3).font = Font(bold=True)
    return _make_response_xlsx(wb, f"BalanceSheet_{fy}.xlsx")

def export_stock_excel(data, fy):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Stock Summary"
    xl_title(ws, f"STOCK SUMMARY — {fy}", 5)
    xl_header(ws, ["Item Name","Unit","In Qty","Out Qty","Closing Qty"], 2)
    for d in data:
        ws.append([d["name"], d["unit"], d["in_qty"], d["out_qty"], d["closing"]])
    return _make_response_xlsx(wb, f"StockSummary_{fy}.xlsx")

def export_outstanding_excel(data, party_type, fy):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = f"{party_type} Outstanding"
    xl_title(ws, f"{party_type.upper()} OUTSTANDING — {fy}", 6)
    xl_header(ws, ["Party Name","GSTIN","Phone","Billed","Paid","Pending"], 2)
    for d in data:
        ws.append([d["name"], d["gstin"], d["phone"], d["billed"], d["paid"], d["pending"]])
    return _make_response_xlsx(wb, f"Outstanding_{party_type}_{fy}.xlsx")

def export_audit_excel(records):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Audit Trail"
    xl_header(ws, ["Date","User","Action","Model","Record ID","IP Address"], 1)
    for r in records:
        ws.append([r.created_at, r.user_id, r.action,
                   r.model_name, r.record_id, r.ip_address])
    return _make_response_xlsx(wb, "AuditTrail.xlsx")

def export_cashflow_excel(operating, investing, financing, net_cash, fy):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Cash Flow"
    xl_title(ws, f"CASH FLOW STATEMENT — {fy}", 2)
    xl_header(ws, ["Activity","Amount (₹)"], 2)
    for row in [
        ["A. Operating Activities", operating],
        ["B. Investing Activities", investing],
        ["C. Financing Activities", financing],
        ["NET CASH FLOW (A+B+C)", net_cash],
    ]: ws.append(row)
    return _make_response_xlsx(wb, f"CashFlow_{fy}.xlsx")

# ═══════════════════════════════════════════
#  PDF EXPORT FUNCTIONS (reportlab)
# ═══════════════════════════════════════════
def _pdf_styles():
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Heading1"],
                                  fontSize=16, spaceAfter=10, textColor=colors.HexColor("#1F4E79"),
                                  alignment=1)
    sub_style   = ParagraphStyle("sub", parent=styles["Normal"],
                                  fontSize=9, textColor=colors.grey, alignment=1)
    return styles, title_style, sub_style

def _pdf_table_style(header_color="#1F4E79"):
    return TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor(header_color)),
        ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
        ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",   (0,0), (-1,0), 9),
        ("FONTSIZE",   (0,1), (-1,-1), 8),
        ("ALIGN",      (0,0), (-1,0), "CENTER"),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#EBF3FB")]),
        ("GRID",       (0,0), (-1,-1), 0.3, colors.grey),
        ("TOPPADDING", (0,0), (-1,-1), 3),
        ("BOTTOMPADDING", (0,0), (-1,-1), 3),
    ])

def export_ledger_pdf(account, lines, opening_dr, opening_cr, fy, start, end):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                             leftMargin=0.5*inch, rightMargin=0.5*inch)
    _, ts, ss = _pdf_styles()
    elements = [
        Paragraph(f"LEDGER ACCOUNT — {account.name if account else ''}", ts),
        Paragraph(f"Financial Year: {fy} | {start} to {end}", ss),
        Spacer(1, 12),
    ]
    table_data = [["Date","Voucher No","Type","Narration","Dr (₹)","Cr (₹)","Balance (₹)"]]
    table_data.append(["Opening","","","",
                        f"{opening_dr:,.2f}", f"{opening_cr:,.2f}",
                        f"{opening_dr-opening_cr:,.2f}"])
    for ln in lines:
        table_data.append([
            str(ln["date"]), ln["voucher_no"], ln["type"],
            (ln["narration"] or "")[:40],
            f"{ln['dr']:,.2f}" if ln["dr"] else "—",
            f"{ln['cr']:,.2f}" if ln["cr"] else "—",
            f"{ln['balance']:,.2f}",
        ])
    t = Table(table_data, colWidths=[70,80,60,220,80,80,90])
    t.setStyle(_pdf_table_style())
    elements.append(t)
    doc.build(elements)
    return _make_response_pdf(buffer, f"Ledger_{account.name if account else 'All'}.pdf")

def export_tb_pdf(data, total_dr, total_cr, fy):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                             leftMargin=0.5*inch, rightMargin=0.5*inch)
    _, ts, ss = _pdf_styles()
    elements = [
        Paragraph(f"TRIAL BALANCE — {fy}", ts),
        Spacer(1, 12),
    ]
    table_data = [["Account Name","Group","Dr (₹)","Cr (₹)","Net Dr","Net Cr"]]
    for d in data:
        table_data.append([d["name"], d["group"],
                           f"{d['dr']:,.2f}", f"{d['cr']:,.2f}",
                           f"{d['net_dr']:,.2f}", f"{d['net_cr']:,.2f}"])
    table_data.append(["TOTAL","",f"{total_dr:,.2f}",f"{total_cr:,.2f}",
                        f"{total_dr:,.2f}",f"{total_cr:,.2f}"])
    t = Table(table_data, colWidths=[180,100,70,70,70,70])
    t.setStyle(_pdf_table_style())
    elements.append(t)
    doc.build(elements)
    return _make_response_pdf(buffer, f"TrialBalance_{fy}.pdf")

def export_pl_pdf(income_data, expense_data, total_income, total_expense, net_profit, fy):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                             leftMargin=0.5*inch, rightMargin=0.5*inch)
    _, ts, ss = _pdf_styles()
    elements = [Paragraph(f"PROFIT & LOSS ACCOUNT — {fy}", ts), Spacer(1, 12)]
    table_data = [["Account Head","Income (₹)","Expense (₹)","Net (₹)"]]
    all_keys = list(set(list(income_data.keys())+list(expense_data.keys())))
    for g in all_keys:
        inc = income_data.get(g,0); exp = expense_data.get(g,0)
        table_data.append([g,
                           f"{inc:,.2f}" if inc else "—",
                           f"{exp:,.2f}" if exp else "—",
                           f"{inc-exp:,.2f}"])
    table_data.append(["TOTAL", f"{total_income:,.2f}",
                        f"{total_expense:,.2f}", f"{net_profit:,.2f}"])
    t = Table(table_data, colWidths=[180,100,100,100])
    t.setStyle(_pdf_table_style("#145A32" if net_profit >= 0 else "#922B21"))
    elements.append(t)
    doc.build(elements)
    return _make_response_pdf(buffer, f"ProfitLoss_{fy}.pdf")

def export_bs_pdf(assets, liabilities, total_assets, total_liabilities, fy):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                             leftMargin=0.5*inch, rightMargin=0.5*inch)
    _, ts, ss = _pdf_styles()
    elements = [Paragraph(f"BALANCE SHEET — {fy}", ts), Spacer(1, 12)]
    lib_rows = list(liabilities.items()); ast_rows = list(assets.items())
    max_r = max(len(lib_rows), len(ast_rows))
    table_data = [["Liabilities","Amount (₹)","Assets","Amount (₹)"]]
    for i in range(max_r):
        lib = lib_rows[i] if i < len(lib_rows) else ("","")
        ast = ast_rows[i] if i < len(ast_rows) else ("","")
        table_data.append([lib[0],f"{lib[1]:,.2f}" if lib[1] else "",
                           ast[0],f"{ast[1]:,.2f}" if ast[1] else ""])
    table_data.append(["TOTAL",f"{total_liabilities:,.2f}",
                        "TOTAL",f"{total_assets:,.2f}"])
    t = Table(table_data, colWidths=[150,100,150,100])
    t.setStyle(_pdf_table_style())
    elements.append(t)
    doc.build(elements)
    return _make_response_pdf(buffer, f"BalanceSheet_{fy}.pdf")

def export_stock_pdf(data, fy):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                             leftMargin=0.5*inch, rightMargin=0.5*inch)
    _, ts, ss = _pdf_styles()
    elements = [Paragraph(f"STOCK SUMMARY — {fy}", ts), Spacer(1, 12)]
    table_data = [["Item Name","Unit","In Qty","Out Qty","Closing Qty","Status"]]
    for d in data:
        table_data.append([d["name"], d["unit"],
                           f"{d['in_qty']:,.3f}", f"{d['out_qty']:,.3f}",
                           f"{d['closing']:,.3f}", d["status"]])
    t = Table(table_data, colWidths=[160,50,70,70,70,60])
    ts2 = _pdf_table_style()
    # Color low-stock rows red
    for i, d in enumerate(data, 1):
        if d["status"] == "Low":
            ts2.add("BACKGROUND", (0,i), (-1,i), colors.HexColor("#FADBD8"))
    t.setStyle(ts2)
    elements.append(t)
    doc.build(elements)
    return _make_response_pdf(buffer, f"StockSummary_{fy}.pdf")

def export_outstanding_pdf(data, party_type, fy):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                             leftMargin=0.5*inch, rightMargin=0.5*inch)
    _, ts, ss = _pdf_styles()
    elements = [Paragraph(f"{party_type.upper()} OUTSTANDING — {fy}", ts), Spacer(1, 12)]
    table_data = [["Party Name","GSTIN","Phone","Billed (₹)","Paid (₹)","Pending (₹)"]]
    for d in data:
        table_data.append([d["name"], d["gstin"] or "—", d["phone"] or "—",
                           f"{d['billed']:,.2f}", f"{d['paid']:,.2f}",
                           f"{d['pending']:,.2f}"])
    t = Table(table_data, colWidths=[160,120,90,90,90,90])
    t.setStyle(_pdf_table_style())
    elements.append(t)
    doc.build(elements)
    return _make_response_pdf(buffer, f"Outstanding_{party_type}_{fy}.pdf")

def export_cashflow_pdf(operating, investing, financing, net_cash, fy):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                             leftMargin=0.5*inch, rightMargin=0.5*inch)
    _, ts, ss = _pdf_styles()
    elements = [Paragraph(f"CASH FLOW STATEMENT — {fy}", ts), Spacer(1, 12)]
    table_data = [
        ["Activity", "Amount (₹)"],
        ["A. Cash from Operating Activities", f"{operating:,.2f}"],
        ["B. Cash from Investing Activities", f"{investing:,.2f}"],
        ["C. Cash from Financing Activities", f"{financing:,.2f}"],
        ["NET CASH FLOW (A+B+C)", f"{net_cash:,.2f}"],
    ]
    t = Table(table_data, colWidths=[300, 120])
    t.setStyle(_pdf_table_style())
    elements.append(t)
    doc.build(elements)
    return _make_response_pdf(buffer, f"CashFlow_{fy}.pdf")
